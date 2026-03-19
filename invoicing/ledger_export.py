from __future__ import annotations

from datetime import date
from typing import TYPE_CHECKING, Callable
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename
from core.app_logging import trace
from utils.billing_date_utils import parse_ymd, today
from utils.outstanding_balance import compute_contract_balance

if TYPE_CHECKING:
    from data.database_service import DatabaseService


@trace
def export_customer_ledger_xlsx(
    parent: tk.Misc,
    db: "DatabaseService",
    customer_id: int,
    customer_name: str,
    customer_phone: str,
    customer_company: str,
    log_action: Callable[[str, str], None],
    default_date: str | None = None,
    # Keep old parameter for backward compat — ignored if present.
    tree: tk.Misc | None = None,
    summary_text: str | None = None,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messagebox.showerror("Missing Dependency", "openpyxl is required for XLSX export.")
        return

    default_date = default_date or date.today().isoformat()
    fp = asksaveasfilename(
        title="Export Ledger",
        defaultextension=".xlsx",
        initialfile=f"ledger_{customer_name.replace(' ', '_')}_{default_date}.xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        parent=parent,
    )
    if not fp:
        return

    # ── Query data from database ────────────────────────────────────
    contracts = db.get_contracts_for_customer_ledger(customer_id)
    as_of = today()
    grand_billed = 0.0
    grand_paid = 0.0

    # ── Build workbook ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"

    bold = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="DDFFDD")
    gray_fill = PatternFill("solid", fgColor="EEEEEE")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    table_cols = ["A", "B", "C", "D", "E", "F", "G"]

    ws.append(["Customer Ledger"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = left

    ws.append(["Customer ID", customer_id, "Name", customer_name])
    ws.append(["Phone", customer_phone or "—", "Company", customer_company or "—"])
    for cell in ws[2]:
        cell.font = bold
    for cell in ws[3]:
        cell.font = bold
    ws.append([])

    header_row = ["Entry", "Date / Period", "Billed", "Paid", "Outstanding", "Scope / Method", "Notes / Reference"]
    ws.append(header_row)
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill

    ws.freeze_panes = "A6"

    # ── Populate rows from database ─────────────────────────────────
    for contract in contracts:
        contract_id_val = int(contract["id"])
        is_active = bool(contract["is_active"])
        rate = float(contract["monthly_rate"])
        scope = contract["scope"]
        truck_info = contract["truck_info"] or ""
        start_d = parse_ymd(contract["start_d"]) if contract["start_d"] else None
        end_d = parse_ymd(contract["end_d"]) if contract["end_d"] else None
        start_raw = contract["start_raw"] or ""
        end_raw = contract["end_raw"] or "ongoing"
        contract_notes = contract["notes"] or ""
        status_lbl = "Active" if is_active else "Inactive"

        payments = db.get_payments_for_contract(contract_id_val)
        total_paid = sum(float(p["amount"]) for p in payments)
        bal = compute_contract_balance(
            monthly_rate=rate,
            start_date_value=contract["start_d"],
            end_date_value=contract["end_d"],
            paid_total=total_paid,
            as_of_date=as_of,
        )
        billed = bal.expected_amount if bal else 0.0
        outstanding = bal.outstanding if bal else 0.0

        grand_billed += billed
        grand_paid += total_paid

        detail_str = f"{scope}" + (f"  {truck_info}" if truck_info else "")
        fill = green_fill if is_active else gray_fill

        # Contract row
        ws.append([
            f"Contract #{contract_id_val}  [{status_lbl}]",
            f"{start_raw} \u2192 {end_raw}",
            billed,
            total_paid,
            outstanding,
            detail_str,
            contract_notes,
        ])
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = bold
        ws.cell(ws.max_row, 3).number_format = "$#,##0.00"
        ws.cell(ws.max_row, 4).number_format = "$#,##0.00"
        ws.cell(ws.max_row, 5).number_format = "$#,##0.00"

        # Payment rows
        if not payments:
            ws.append(["(no payments)", "", "", "", "", "", ""])
        else:
            for idx, payment in enumerate(payments, start=1):
                ref_notes = " | ".join(filter(None, [payment["reference"], payment["notes"]]))
                amt = float(payment["amount"])
                ws.append([
                    f"Payment #{idx}",
                    payment["paid_at"],
                    None,
                    amt,
                    None,
                    payment["method"],
                    ref_notes,
                ])
                ws.cell(ws.max_row, 4).number_format = "$#,##0.00"

    # ── Summary row ─────────────────────────────────────────────────
    ws.append([])
    grand_outstanding = grand_billed - grand_paid
    computed_summary = (
        f"Contracts: {len(contracts)}     "
        f"Total Billed: ${grand_billed:.2f}     "
        f"Total Paid: ${grand_paid:.2f}     "
        f"Total Outstanding: ${grand_outstanding:.2f}"
    )
    ws.append([computed_summary])
    ws.merge_cells(f"A{ws.max_row}:G{ws.max_row}")
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 1).alignment = left

    for col in table_cols:
        max_len = 0
        for row_i in range(1, ws.max_row + 1):
            val = ws[f"{col}{row_i}"].value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[col].width = min(max(max_len + 2, 12), 48)

    try:
        wb.save(fp)
    except PermissionError:
        messagebox.showerror(
            "Export Failed",
            f"Cannot save to:\n{fp}\n\nThe file may be open in another program. "
            "Please close it and try again.",
        )
        return
    except Exception as exc:
        messagebox.showerror("Export Failed", f"Failed to save ledger:\n{exc}")
        return
    log_action("EXPORT_LEDGER", f"Exported ledger for Customer ID {customer_id} ({customer_name}) to {fp}")
    messagebox.showinfo("Exported", f"Ledger saved to:\n{fp}")
